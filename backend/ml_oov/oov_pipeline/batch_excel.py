from __future__ import annotations

import argparse
from typing import Optional

from openpyxl import load_workbook

from .adapters import make_mloov_predictor
from .lexicon_adapter import analyzer_lexicon_matcher
from .pipeline import run_oov_pipeline


def _detect_text_col(headers) -> Optional[int]:
    """
    헤더에서 '부정경험 글' 컬럼을 최대한 자동으로 찾는다.
    - 정확히 일치하면 최우선
    - 아니면 키워드 점수로 선택
    """
    if not headers:
        return None

    headers_s = [("" if h is None else str(h).strip()) for h in headers]

    # 1) 정확히 일치 후보(너희 데이터에 맞춰 우선순위)
    exact_candidates = [
        "부정경험 글",
        "부정경험글",
        "부정 경험 글",
        "일기",
        "내용",
        "텍스트",
    ]
    for name in exact_candidates:
        for i, h in enumerate(headers_s):
            if h == name:
                return i

    # 2) 키워드 점수로 탐지
    keys = ["부정", "경험", "글", "일기", "내용", "텍스트"]
    best_i, best_score = None, 0
    for i, h in enumerate(headers_s):
        if not h:
            continue
        score = sum(k in h for k in keys)
        if score > best_score:
            best_score = score
            best_i = i

    return best_i if best_score >= 2 else None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="입력 엑셀 파일 경로")
    ap.add_argument("--sheet", default=None, help="시트명(미지정시 첫 시트)")
    ap.add_argument("--text_col", default=None, help="텍스트 컬럼명(미지정시 자동 탐지)")
    ap.add_argument("--ckpt", default="models/koelectra_oov", help="A 모델 체크포인트 디렉토리")
    ap.add_argument("--threshold", type=float, default=0.0, help="토큰 후보 confidence threshold(보통 0.0)")
    ap.add_argument("--max_rows", type=int, default=0, help="0이면 전체, 아니면 앞에서부터 N행만")
    ap.add_argument("--no_write", action="store_true", help="후보 파일 저장 비활성화(테스트용)")
    args = ap.parse_args()

    # A predictor
    predictor = make_mloov_predictor(ckpt_dir=args.ckpt, device="cpu")

    # Lexicon matcher: 기존 NegativeEmotionAnalyzer 기반 (활용형+겹침방지)
    lexicon_matcher = analyzer_lexicon_matcher

    wb = load_workbook(args.excel, data_only=True)
    sheet = args.sheet
    if sheet is None:
        ws = wb[wb.sheetnames[0]]
    else:
        # 숫자로 들어오면 인덱스로, 아니면 이름으로
        if str(sheet).isdigit():
            ws = wb.worksheets[int(sheet)]
        else:
            ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("[batch_excel] empty sheet")
        return

    headers = rows[0]
    headers_s = [("" if h is None else str(h).strip()) for h in headers]

    if args.text_col:
        idx = None
        for i, h in enumerate(headers_s):
            if h == args.text_col:
                idx = i
                break
        if idx is None:
            raise ValueError(f"[batch_excel] text_col '{args.text_col}' not found. headers={headers_s}")
    else:
        idx = _detect_text_col(headers_s)
        if idx is None:
            raise ValueError(f"[batch_excel] could not auto-detect text column. headers={headers_s}")

    print(f"[batch_excel] sheet='{ws.title}' text_col_index={idx} header='{headers_s[idx]}'")

    overrides = {
        "conf_threshold": float(args.threshold),
        "write_candidates": (not args.no_write),
        "enabled": True,
    }

    processed = 0
    for r in rows[1:]:
        if args.max_rows and processed >= args.max_rows:
            break

        if idx >= len(r):
            continue

        text = r[idx]
        if text is None:
            continue

        text = str(text).strip()
        if not text:
            continue

        run_oov_pipeline(
            text,
            predictor=predictor,
            lexicon_matcher=lexicon_matcher,
            config_overrides=overrides,
        )

        processed += 1
        if processed % 50 == 0:
            print(f"[batch_excel] processed={processed}")

    print(f"[batch_excel] done. processed={processed}")


if __name__ == "__main__":
    main()
