# backend/ml_oov/make_gold_from_report.py
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook

def main(
    xlsx_path: str = "backend/reports/oov/report.xlsx",
    out_path: str = "backend/data/oov_gold/train.jsonl",
    sheet: str = None,
    key_col: str = "A",
    example_col: str = "E",
    keep_col: str = "F",
    label: str = "NEG_OOV",
):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet] if sheet else wb.active

    out_lines = []
    n_keep = 0
    n_fail = 0

    # 1행은 헤더라고 가정
    for r in range(2, ws.max_row + 1):
        key = ws[f"{key_col}{r}"].value
        ex = ws[f"{example_col}{r}"].value
        keep = ws[f"{keep_col}{r}"].value

        if not key or not ex:
            continue
        if str(keep).strip() not in ("1", "True", "true", "Y", "y"):
            continue

        n_keep += 1
        text = str(ex)

        # key가 문장에 여러 번 나오면 첫 번째만
        start = text.find(str(key))
        if start < 0:
            n_fail += 1
            continue
        end = start + len(str(key))

        row = {
            "id": f"r{r}",
            "text": text,
            "oov_spans": [{"start": start, "end": end, "label": label}],
        }
        out_lines.append(json.dumps(row, ensure_ascii=False))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    Path(out_path).write_text("\n".join(out_lines) + ("\n" if out_lines else ""), encoding="utf-8")

    print(f"[make_gold] keep_rows={n_keep} wrote={len(out_lines)} fail_no_match={n_fail} -> {out_path}")

if __name__ == "__main__":
    main()
